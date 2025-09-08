import boto3
from botocore.exceptions import ClientError
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone, timedelta
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO

# -------------------------
# CONFIG
# -------------------------
S3_BUCKET = "audit-log-sm20-bucket"
S3_KEY_PREFIX = "status/Sapphire-PRD"
LOOKBACK_MINUTES = 60
REGION = "ap-south-1"

# Account Details
ACCOUNT_NAME = "Sapphire-PRD"

# Allowed Linux paths
ALLOWED_DISK_PATHS = ['/', '/usr', '/hana']

# Boto3 clients
cloudwatch = boto3.client('cloudwatch', region_name=REGION)
ec2 = boto3.client('ec2', region_name=REGION)
s3 = boto3.client('s3', region_name=REGION)
sts = boto3.client('sts', region_name=REGION)  # Added for account number

# -------------------------
# Helpers
# -------------------------
def get_account_info():
    """Get AWS account number"""
    try:
        response = sts.get_caller_identity()
        account_number = response.get('Account', 'Unknown')
        return account_number
    except ClientError:
        return 'Unknown'

def get_running_instances():
    response = ec2.describe_instances(
        Filters=[{'Name': 'instance-state-name', 'Values': ['running']}]
    )
    instances = {}
    for reservation in response['Reservations']:
        for instance in reservation['Instances']:
            instance_id = instance['InstanceId']
            name = instance_id
            os_type = instance.get("PlatformDetails", "Linux/UNIX")
            instance_type = instance.get('InstanceType', 'Unknown')
            state = instance.get('State', {}).get('Name', 'unknown')
            for tag in instance.get('Tags', []):
                if tag['Key'] == 'Name':
                    name = tag['Value']
                    break
            instances[instance_id] = {
                "name": name, 
                "os_type": os_type, 
                "instance_type": instance_type,
                "state": state
            }
    return instances

def get_all_instances():
    """Get all instances (running and stopped) with their details"""
    response = ec2.describe_instances()
    instances = {}
    for reservation in response['Reservations']:
        for instance in reservation['Instances']:
            instance_id = instance['InstanceId']
            name = instance_id
            os_type = instance.get("PlatformDetails", "Linux/UNIX")
            instance_type = instance.get('InstanceType', 'Unknown')
            state = instance.get('State', {}).get('Name', 'unknown')
            for tag in instance.get('Tags', []):
                if tag['Key'] == 'Name':
                    name = tag['Value']
                    break
            instances[instance_id] = {
                "name": name, 
                "os_type": os_type, 
                "instance_type": instance_type,
                "state": state
            }
    return instances

def get_instance_type_specs(instance_type):
    """Get vCPU and memory specifications for instance type"""
    try:
        response = ec2.describe_instance_types(InstanceTypes=[instance_type])
        if response['InstanceTypes']:
            spec = response['InstanceTypes'][0]
            vcpu_count = spec.get('VCpuInfo', {}).get('DefaultVCpus', 'Unknown')
            memory_mb = spec.get('MemoryInfo', {}).get('SizeInMiB', 0)
            memory_gb = round(memory_mb / 1024, 1) if memory_mb else 'Unknown'
            return vcpu_count, memory_gb
    except ClientError:
        pass
    return 'Unknown', 'Unknown'

def get_current_metric_value(metric_name, namespace, dimensions):
    """Get the most recent (current) metric value"""
    end_time = datetime.now(timezone.utc)
    start_time = end_time - timedelta(minutes=10)  # Look at last 10 minutes for current value
    try:
        response = cloudwatch.get_metric_data(
            MetricDataQueries=[
                {
                    'Id': 'm1',
                    'MetricStat': {
                        'Metric': {
                            'Namespace': namespace,
                            'MetricName': metric_name,
                            'Dimensions': dimensions
                        },
                        'Period': 300,
                        'Stat': 'Average'  # Use Average for more realistic current values
                    },
                    'ReturnData': True
                }
            ],
            StartTime=start_time,
            EndTime=end_time
        )
        values = response['MetricDataResults'][0]['Values']
        return values[-1] if values else None  # Return the most recent value
    except ClientError as e:
        print(f"Error getting metric {metric_name}: {e}")
        return None

def get_linux_disk_dimensions(instance_id):
    """Linux disks: only /, /usr/, /hana"""
    dimensions_list = []
    try:
        paginator = cloudwatch.get_paginator('list_metrics')
        for page in paginator.paginate(
            Namespace='CWAgent',
            MetricName='disk_used_percent',
            Dimensions=[{'Name': 'InstanceId', 'Value': instance_id}]
        ):
            for metric in page['Metrics']:
                path = next((d['Value'] for d in metric['Dimensions'] if d['Name'] == 'path'), None)
                if path and (path == '/' or path.startswith('/usr') or path.startswith('/hana')):
                    dimensions_list.append(metric['Dimensions'])
    except ClientError:
        pass
    return dimensions_list

def get_windows_disks_all(instance_id):
    """Windows disks: fetch all drives (C:, D:, etc.)"""
    disks = []
    processed_drives = set()  # Track processed drives to avoid duplicates
    
    try:
        print(f"DEBUG: Fetching Windows disk metrics for instance {instance_id}")
        
        # List all LogicalDisk % Free Space metrics for this instance
        paginator = cloudwatch.get_paginator('list_metrics')
        for page in paginator.paginate(
            Namespace='CWAgent',
            MetricName='LogicalDisk % Free Space',
            Dimensions=[{'Name': 'InstanceId', 'Value': instance_id}]
        ):
            for metric in page['Metrics']:
                print(f"DEBUG: Found metric with dimensions: {metric['Dimensions']}")
                
                # Extract the instance dimension which contains the drive letter for Windows
                drive_letter = None
                objectname = None
                
                for dim in metric['Dimensions']:
                    if dim['Name'] == 'instance':
                        # The instance dimension contains the drive letter (C:, D:, etc.)
                        instance_value = dim['Value']
                        # Check if this looks like a drive letter
                        if instance_value and ':' in instance_value and instance_value != '_Total':
                            drive_letter = instance_value
                    elif dim['Name'] == 'objectname':
                        objectname = dim['Value']
                
                print(f"DEBUG: Drive: {drive_letter}, ObjectName: {objectname}")
                
                # Process valid drive letters that haven't been processed yet
                if drive_letter and drive_letter not in processed_drives:
                    if objectname == 'LogicalDisk':  # Ensure it's a LogicalDisk metric
                        print(f"DEBUG: Processing drive {drive_letter}")
                        processed_drives.add(drive_letter)
                        
                        # Get the free space percentage
                        free_percent = get_current_metric_value('LogicalDisk % Free Space', 'CWAgent', metric['Dimensions'])
                        if free_percent is not None:
                            used_percent = 100 - free_percent
                            print(f"DEBUG: Drive {drive_letter} - Free: {free_percent:.1f}%, Used: {used_percent:.1f}%")
                            disks.append({'path': drive_letter, 'used_percent': used_percent})
                        else:
                            print(f"DEBUG: No data available for drive {drive_letter}")
                            disks.append({'path': drive_letter, 'used_percent': 'NA'})
    
    except ClientError as e:
        print(f"ERROR: Failed to fetch Windows disk metrics: {e}")
    
    # Sort disks by drive letter for consistent ordering
    disks.sort(key=lambda x: x['path'])
    
    print(f"DEBUG: Found {len(disks)} Windows disks: {[d['path'] for d in disks]}")
    return disks

def get_memory_metric(instance_id):
    """Memory % utilization for both Linux and Windows"""
    print(f"DEBUG: Fetching memory metrics for instance {instance_id}")
    
    # Try Linux memory metric first
    try:
        paginator = cloudwatch.get_paginator('list_metrics')
        for page in paginator.paginate(
            Namespace='CWAgent',
            MetricName='mem_used_percent',
            Dimensions=[{'Name': 'InstanceId', 'Value': instance_id}]
        ):
            for metric in page['Metrics']:
                print(f"DEBUG: Found Linux memory metric: {metric['Dimensions']}")
                val = get_current_metric_value('mem_used_percent', 'CWAgent', metric['Dimensions'])
                if val is not None:
                    print(f"DEBUG: Linux memory usage: {val:.1f}%")
                    return {'used_percent': val}
    except ClientError as e:
        print(f"DEBUG: Linux memory metric not found: {e}")
    
    # Try Windows memory metrics
    try:
        print("DEBUG: Trying Windows memory metrics...")
        
        # List all Memory % Committed Bytes In Use metrics
        for page in paginator.paginate(
            Namespace='CWAgent',
            MetricName='Memory % Committed Bytes In Use',
            Dimensions=[{'Name': 'InstanceId', 'Value': instance_id}]
        ):
            for metric in page['Metrics']:
                print(f"DEBUG: Found Windows memory metric with dimensions: {metric['Dimensions']}")
                
                # Check if objectname is Memory
                objectname = None
                for dim in metric['Dimensions']:
                    if dim['Name'] == 'objectname':
                        objectname = dim['Value']
                
                print(f"DEBUG: Memory objectname: {objectname}")
                
                # Process only Memory objectname metrics
                if objectname == 'Memory':
                    val = get_current_metric_value('Memory % Committed Bytes In Use', 'CWAgent', metric['Dimensions'])
                    if val is not None:
                        print(f"DEBUG: Windows memory usage: {val:.1f}%")
                        return {'used_percent': val}
                    else:
                        print("DEBUG: No memory data available")
    
    except ClientError as e:
        print(f"ERROR: Windows memory metric failed: {e}")
    
    # Try alternative Windows memory metric - Memory Available Mbytes
    try:
        print("DEBUG: Trying alternative Windows memory metric (Memory Available Mbytes)...")
        
        for page in paginator.paginate(
            Namespace='CWAgent',
            MetricName='Memory Available Mbytes',
            Dimensions=[{'Name': 'InstanceId', 'Value': instance_id}]
        ):
            for metric in page['Metrics']:
                print(f"DEBUG: Found Windows Memory Available Mbytes metric: {metric['Dimensions']}")
                
                val = get_current_metric_value('Memory Available Mbytes', 'CWAgent', metric['Dimensions'])
                if val is not None:
                    # Get instance type to calculate percentage
                    # This is a rough estimate - you may need to get total memory another way
                    print(f"DEBUG: Windows memory available: {val} MB")
                    # For now, return NA since we can't calculate percentage without total memory
                    return {'used_percent': 'NA', 'note': f'Available: {val:.0f} MB'}
    
    except ClientError as e:
        print(f"ERROR: Alternative Windows memory metric failed: {e}")
    
    print("DEBUG: No memory metrics found - returning NA")
    return {'used_percent': 'NA'}

def get_instance_metrics(instance_id):
    print(f"DEBUG: Starting metrics collection for instance {instance_id}")
    metrics = {'instance_id': instance_id, 'cpu': None, 'memory': {}, 'disks': []}
    
    try:
        # CPU - using current value
        print(f"DEBUG: Fetching CPU metrics for {instance_id}")
        metrics['cpu'] = get_current_metric_value(
            'CPUUtilization', 'AWS/EC2',
            [{'Name': 'InstanceId', 'Value': instance_id}]
        )
        if metrics['cpu'] is not None:
            print(f"DEBUG: CPU usage: {metrics['cpu']:.1f}%")
        else:
            print(f"DEBUG: No CPU data available")
        
        # Memory
        metrics['memory'] = get_memory_metric(instance_id)
        
        # Check if Windows or Linux based on existing metrics
        is_windows = False
        
        # Quick check for Windows by looking for LogicalDisk metrics
        try:
            paginator = cloudwatch.get_paginator('list_metrics')
            for page in paginator.paginate(
                Namespace='CWAgent',
                MetricName='LogicalDisk % Free Space',
                Dimensions=[{'Name': 'InstanceId', 'Value': instance_id}]
            ):
                if page['Metrics']:
                    is_windows = True
                    break
        except:
            pass
        
        if is_windows:
            print(f"DEBUG: Instance {instance_id} detected as Windows")
            # Windows disks
            windows_disks = get_windows_disks_all(instance_id)
            metrics['disks'].extend(windows_disks)
        else:
            print(f"DEBUG: Instance {instance_id} detected as Linux")
            # Linux disks
            linux_disks = get_linux_disk_dimensions(instance_id)
            print(f"DEBUG: Found {len(linux_disks)} Linux disk configurations")
            for dims in linux_disks:
                path = next((d['Value'] for d in dims if d['Name'] == 'path'), 'N/A')
                val = get_current_metric_value('disk_used_percent', 'CWAgent', dims)
                metrics['disks'].append({'path': path, 'used_percent': val if val is not None else 'NA'})
        
        print(f"DEBUG: Total disks found: {len(metrics['disks'])}")
        
    except Exception as e:
        print(f"ERROR: Failed to get metrics for {instance_id}: {str(e)}")
        import traceback
        print(f"DEBUG: Full traceback: {traceback.format_exc()}")
        metrics['error'] = f"Failed to get metrics: {str(e)}"
    
    return metrics

def apply_conditional_formatting(cell, used_percent):
    """Apply conditional formatting based on usage percentage"""
    if isinstance(used_percent, (int, float)):
        if used_percent > 60:
            # Red background for > 60%
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        elif used_percent > 50:
            # Yellow background for > 50%
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def apply_table_style(sheet, start_row, end_row, max_col):
    """Add borders, alignment, autosize"""
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, max_col + 1):
        sheet.column_dimensions[get_column_letter(col)].auto_size = True

def create_summary_section(sheet, all_instances, metrics_data, account_number):
    """Create summary section at the beginning of the report"""
    row_num = 1
    
    # Title
    title_cell = sheet.cell(row=row_num, column=1, value="EC2 Instance Metrics Summary Report")
    title_cell.font = Font(bold=True, size=16)
    sheet.merge_cells(f'A{row_num}:D{row_num}')
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    row_num += 2
    
    # Account Information Section
    sheet.cell(row=row_num, column=1, value="Account Name:").font = Font(bold=True)
    sheet.cell(row=row_num, column=2, value=ACCOUNT_NAME)
    row_num += 1
    
    sheet.cell(row=row_num, column=1, value="Account Number:").font = Font(bold=True)
    sheet.cell(row=row_num, column=2, value=account_number)
    row_num += 1
    
    sheet.cell(row=row_num, column=1, value="Region:").font = Font(bold=True)
    sheet.cell(row=row_num, column=2, value=REGION)
    row_num += 2
    
    # Instance Types Summary Table
    thin = Side(border_style="thin", color="000000")
    thick = Side(border_style="thick", color="000000")
    
    # Header
    sheet.cell(row=row_num, column=1, value="Instance Types").font = Font(bold=True)
    sheet.cell(row=row_num, column=2, value="Running").font = Font(bold=True)
    sheet.cell(row=row_num, column=3, value="Stopped").font = Font(bold=True)
    sheet.cell(row=row_num, column=4, value="Total").font = Font(bold=True)
    
    # Apply borders to header
    for col in range(1, 5):
        cell = sheet.cell(row=row_num, column=col)
        cell.border = Border(top=thick, left=thin, right=thin, bottom=thick)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row_num += 1
    
    # Count instances by type and state
    windows_running = 0
    windows_stopped = 0
    linux_running = 0
    linux_stopped = 0
    suse_running = 0
    suse_stopped = 0
    
    for instance_id, info in all_instances.items():
        os_type = info.get("os_type", "Unknown")
        state = info.get("state", "unknown")
        
        if "Windows" in os_type:
            if state == "running":
                windows_running += 1
            else:
                windows_stopped += 1
        elif "SUSE" in os_type:
            if state == "running":
                suse_running += 1
            else:
                suse_stopped += 1
        else:  # Linux/UNIX
            if state == "running":
                linux_running += 1
            else:
                linux_stopped += 1
    
    # Windows row
    sheet.cell(row=row_num, column=1, value="Windows")
    sheet.cell(row=row_num, column=2, value=windows_running)
    sheet.cell(row=row_num, column=3, value=windows_stopped)
    sheet.cell(row=row_num, column=4, value=windows_running + windows_stopped)
    for col in range(1, 5):
        cell = sheet.cell(row=row_num, column=col)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row_num += 1
    
    # Linux/UNIX row
    sheet.cell(row=row_num, column=1, value="Linux/UNIX")
    sheet.cell(row=row_num, column=2, value=linux_running)
    sheet.cell(row=row_num, column=3, value=linux_stopped)
    sheet.cell(row=row_num, column=4, value=linux_running + linux_stopped)
    for col in range(1, 5):
        cell = sheet.cell(row=row_num, column=col)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row_num += 1
    
    # SUSE Linux row
    sheet.cell(row=row_num, column=1, value="SUSE Linux")
    sheet.cell(row=row_num, column=2, value=suse_running)
    sheet.cell(row=row_num, column=3, value=suse_stopped)
    sheet.cell(row=row_num, column=4, value=suse_running + suse_stopped)
    for col in range(1, 5):
        cell = sheet.cell(row=row_num, column=col)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row_num += 1
    
    # Total Servers row
    total_running = windows_running + linux_running + suse_running
    total_stopped = windows_stopped + linux_stopped + suse_stopped
    total_cell = sheet.cell(row=row_num, column=1, value="Total Servers")
    total_cell.font = Font(bold=True)
    sheet.cell(row=row_num, column=2, value=total_running).font = Font(bold=True)
    sheet.cell(row=row_num, column=3, value=total_stopped).font = Font(bold=True)
    sheet.cell(row=row_num, column=4, value=total_running + total_stopped).font = Font(bold=True)
    for col in range(1, 5):
        cell = sheet.cell(row=row_num, column=col)
        cell.border = Border(top=thick, left=thin, right=thin, bottom=thick)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row_num += 3
    
    # Collect alerts for running instances
    yellow_alerts = []  # Warning alerts
    red_alerts = []     # Action required alerts
    
    print(f"DEBUG: Processing {len(metrics_data)} instances for alerts")
    
    for data in metrics_data:
        if 'error' not in data:
            instance_id = data['instance_id']
            instance_info = all_instances.get(instance_id, {})
            instance_name = instance_info.get('name', instance_id)
            
            print(f"DEBUG: Checking alerts for {instance_name} ({instance_id})")
            
            # Check CPU - lowered threshold for testing
            cpu = data.get('cpu')
            if isinstance(cpu, (int, float)):
                print(f"DEBUG: CPU usage: {cpu:.1f}%")
                if cpu >= 50:  # Lowered from 85 for testing
                    if cpu > 60:  # Lowered from 95 for testing
                        red_alerts.append({'name': instance_name, 'metric': 'CPU', 'utilization': f'{cpu:.1f}%'})
                        print(f"DEBUG: Added RED alert for CPU: {cpu:.1f}%")
                    else:
                        yellow_alerts.append({'name': instance_name, 'metric': 'CPU', 'utilization': f'{cpu:.1f}%'})
                        print(f"DEBUG: Added YELLOW alert for CPU: {cpu:.1f}%")
            
            # Check Memory - lowered threshold for testing
            mem = data.get('memory', {})
            mem_used = mem.get('used_percent')
            if isinstance(mem_used, (int, float)):
                print(f"DEBUG: Memory usage: {mem_used:.1f}%")
                if mem_used >= 50:  # Lowered from 85 for testing
                    if mem_used > 60:  # Lowered from 95 for testing
                        red_alerts.append({'name': instance_name, 'metric': 'Memory', 'utilization': f'{mem_used:.1f}%'})
                        print(f"DEBUG: Added RED alert for Memory: {mem_used:.1f}%")
                    else:
                        yellow_alerts.append({'name': instance_name, 'metric': 'Memory', 'utilization': f'{mem_used:.1f}%'})
                        print(f"DEBUG: Added YELLOW alert for Memory: {mem_used:.1f}%")
            
            # Check Disks - lowered threshold for testing
            for disk in data.get('disks', []):
                disk_used = disk.get('used_percent')
                if isinstance(disk_used, (int, float)):
                    print(f"DEBUG: Disk {disk.get('path', 'N/A')} usage: {disk_used:.1f}%")
                    if disk_used >= 50:  # Lowered from 85 for testing
                        metric_name = f'Disk ({disk.get("path", "N/A")})'
                        if disk_used > 60:  # Lowered from 90 for testing
                            red_alerts.append({'name': instance_name, 'metric': metric_name, 'utilization': f'{disk_used:.1f}%'})
                            print(f"DEBUG: Added RED alert for {metric_name}: {disk_used:.1f}%")
                        else:
                            yellow_alerts.append({'name': instance_name, 'metric': metric_name, 'utilization': f'{disk_used:.1f}%'})
                            print(f"DEBUG: Added YELLOW alert for {metric_name}: {disk_used:.1f}%")
    
    print(f"DEBUG: Found {len(yellow_alerts)} yellow alerts and {len(red_alerts)} red alerts")
    
    # Yellow Alerts Table (Warning)
    if yellow_alerts:
        # Table Header
        header_cell = sheet.cell(row=row_num, column=1, value="Warning")
        header_cell.font = Font(bold=True, color="000000")
        header_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        sheet.merge_cells(f'A{row_num}:C{row_num}')
        for col in range(1, 4):
            sheet.cell(row=row_num, column=col).border = Border(top=thick, left=thin, right=thin, bottom=thick)
        row_num += 1
        
        # Column Headers
        sheet.cell(row=row_num, column=1, value="Instance Name").font = Font(bold=True)
        sheet.cell(row=row_num, column=2, value="Metric").font = Font(bold=True)
        sheet.cell(row=row_num, column=3, value="Utilization").font = Font(bold=True)
        for col in range(1, 4):
            cell = sheet.cell(row=row_num, column=col)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thick)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row_num += 1
        
        # Data rows
        for alert in yellow_alerts:
            sheet.cell(row=row_num, column=1, value=alert['name'])
            sheet.cell(row=row_num, column=2, value=alert['metric'])
            sheet.cell(row=row_num, column=3, value=alert['utilization'])
            for col in range(1, 4):
                cell = sheet.cell(row=row_num, column=col)
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            row_num += 1
        row_num += 2
    
    # Red Alerts Table (Action Required)
    if red_alerts:
        # Table Header
        header_cell = sheet.cell(row=row_num, column=1, value="Action Required")
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        sheet.merge_cells(f'A{row_num}:C{row_num}')
        for col in range(1, 4):
            sheet.cell(row=row_num, column=col).border = Border(top=thick, left=thin, right=thin, bottom=thick)
        row_num += 1
        
        # Column Headers
        sheet.cell(row=row_num, column=1, value="Instance Name").font = Font(bold=True)
        sheet.cell(row=row_num, column=2, value="Metric").font = Font(bold=True)
        sheet.cell(row=row_num, column=3, value="Utilization").font = Font(bold=True)
        for col in range(1, 4):
            cell = sheet.cell(row=row_num, column=col)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thick)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row_num += 1
        
        # Data rows
        for alert in red_alerts:
            sheet.cell(row=row_num, column=1, value=alert['name'])
            sheet.cell(row=row_num, column=2, value=alert['metric'])
            sheet.cell(row=row_num, column=3, value=alert['utilization'])
            for col in range(1, 4):
                cell = sheet.cell(row=row_num, column=col)
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            row_num += 1
        row_num += 2
    
    # Add separator
    row_num += 1
    separator = sheet.cell(row=row_num, column=1, value="=" * 50)
    sheet.merge_cells(f'A{row_num}:D{row_num}')
    separator.alignment = Alignment(horizontal="center")
    row_num += 1
    
    detail_header = sheet.cell(row=row_num, column=1, value="DETAILED INSTANCE METRICS")
    detail_header.font = Font(bold=True, size=14)
    sheet.merge_cells(f'A{row_num}:D{row_num}')
    detail_header.alignment = Alignment(horizontal="center")
    row_num += 1
    
    separator = sheet.cell(row=row_num, column=1, value="=" * 50)
    sheet.merge_cells(f'A{row_num}:D{row_num}')
    separator.alignment = Alignment(horizontal="center")
    row_num += 2
    
    return row_num

def create_excel_report(metrics_data, instance_meta, all_instances, account_number):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Sapphire-PRD"  # Set sheet name to Sapphire-PRD
    headers = ["Metric Type", "Path", "Used % (Current)", "Free % (Current)"]
    
    # Create summary section first
    row_num = create_summary_section(sheet, all_instances, metrics_data, account_number)
    
    # Process all instances (running and stopped) - Original report continues here
    for instance_id, instance_info in all_instances.items():
        iname = instance_info.get("name", instance_id)
        os_type = instance_info.get("os_type", "Unknown")
        instance_type = instance_info.get("instance_type", "Unknown")
        state = instance_info.get("state", "unknown")
        
        # Get instance type specifications
        vcpu_count, memory_gb = get_instance_type_specs(instance_type)

        sheet.cell(row=row_num, column=1, value=f"Instance ID: {instance_id}").font = Font(bold=True, size=14)
        row_num += 1
        sheet.cell(row=row_num, column=1, value=f"Instance Name: {iname}").font = Font(bold=True)
        row_num += 1
        sheet.cell(row=row_num, column=1, value=f"OS Type: {os_type}").font = Font(bold=True)
        row_num += 1
        sheet.cell(row=row_num, column=1, value=f"Instance Type: {instance_type}").font = Font(bold=True)
        row_num += 1
        sheet.cell(row=row_num, column=1, value=f"vCPU Count: {vcpu_count}").font = Font(bold=True)
        row_num += 1
        sheet.cell(row=row_num, column=1, value=f"Memory: {memory_gb} GB").font = Font(bold=True)
        row_num += 1
        sheet.cell(row=row_num, column=1, value=f"State: {state}").font = Font(bold=True)
        row_num += 2

        # Only show metrics for running instances
        if state == 'running':
            # Find metrics data for this instance
            data = next((d for d in metrics_data if d['instance_id'] == instance_id), None)
            
            if data and 'error' not in data:
                for col_num, header in enumerate(headers, 1):
                    sheet.cell(row=row_num, column=col_num, value=header).font = Font(bold=True)
                header_row = row_num
                row_num += 1

                # CPU row (Used + Free %)
                cpu = data.get('cpu', 'NA')
                sheet.cell(row=row_num, column=1, value="CPU")
                used_cell = sheet.cell(row=row_num, column=3)
                if isinstance(cpu, (int, float)):
                    used = round(cpu, 1)  # Keep one decimal for more precision
                    used_cell.value = used
                    sheet.cell(row=row_num, column=4, value=round(100 - used, 1))
                    # Apply conditional formatting to used % cell
                    apply_conditional_formatting(used_cell, used)
                else:
                    used_cell.value = "NA"
                    sheet.cell(row=row_num, column=4, value="NA")
                row_num += 1

                # Memory row
                mem = data.get('memory', {})
                sheet.cell(row=row_num, column=1, value="Memory")
                used_cell = sheet.cell(row=row_num, column=3)
                if isinstance(mem.get('used_percent'), (int, float)):
                    used = round(mem['used_percent'], 1)
                    used_cell.value = used
                    sheet.cell(row=row_num, column=4, value=round(100 - used, 1))
                    # Apply conditional formatting to used % cell
                    apply_conditional_formatting(used_cell, used)
                else:
                    used_cell.value = "NA"
                    sheet.cell(row=row_num, column=4, value="NA")
                    # Add note if available
                    if 'note' in mem:
                        sheet.cell(row=row_num, column=2, value=mem['note'])
                row_num += 1

                # Disk rows
                disks = data.get('disks', [])
                if disks:
                    for disk in disks:
                        sheet.cell(row=row_num, column=1, value="Disk")
                        sheet.cell(row=row_num, column=2, value=disk.get('path', 'N/A'))
                        used_cell = sheet.cell(row=row_num, column=3)
                        if isinstance(disk.get('used_percent'), (int, float)):
                            used = round(disk['used_percent'], 1)
                            used_cell.value = used
                            sheet.cell(row=row_num, column=4, value=round(100 - used, 1))
                            # Apply conditional formatting to used % cell
                            apply_conditional_formatting(used_cell, used)
                        else:
                            used_cell.value = "NA"
                            sheet.cell(row=row_num, column=4, value="NA")
                        row_num += 1
                else:
                    # No disks found
                    sheet.cell(row=row_num, column=1, value="Disk")
                    sheet.cell(row=row_num, column=2, value="No disk metrics available")
                    sheet.cell(row=row_num, column=3, value="NA")
                    sheet.cell(row=row_num, column=4, value="NA")
                    row_num += 1

                apply_table_style(sheet, header_row, row_num - 1, len(headers))
            else:
                # Show error or no metrics available
                error_msg = data.get('error', 'No metrics data') if data else 'No metrics data'
                sheet.cell(row=row_num, column=1, value=f"Metrics: Not available ({error_msg})")
                row_num += 1
        else:
            # For stopped instances
            sheet.cell(row=row_num, column=1, value="Metrics: Not available (instance stopped)")
            row_num += 1
            
        row_num += 2

    # Convert UTC to IST for report generation time
    ist_offset = timedelta(hours=5, minutes=30)  # IST is UTC+5:30
    current_time_ist = datetime.now(timezone.utc) + ist_offset
    
    sheet.cell(row=row_num, column=1, value="Report Generated At").font = Font(bold=True)
    sheet.cell(row=row_num, column=2, value=current_time_ist.strftime("%Y-%m-%d %H:%M:%S IST"))

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def lambda_handler(event, context):
    try:
        print("DEBUG: Starting Lambda execution...")
        
        # Get account information
        account_number = get_account_info()
        print(f"DEBUG: Account Number: {account_number}")
        
        # Get all instances (running and stopped)
        all_instances = get_all_instances()
        print(f"DEBUG: Found {len(all_instances)} total instances")
        
        # Get running instances for metrics collection
        running_instances = get_running_instances()
        print(f"DEBUG: Found {len(running_instances)} running instances")
        
        # Collect metrics only for running instances
        metrics_data = []
        with ThreadPoolExecutor(max_workers=10) as executor:
            future_to_instance = {executor.submit(get_instance_metrics, iid): iid for iid in running_instances}
            for future in as_completed(future_to_instance):
                iid = future_to_instance[future]
                try:
                    result = future.result()
                    metrics_data.append(result)
                    print(f"DEBUG: Successfully collected metrics for {iid}")
                except Exception as e:
                    error_msg = f"Failed: {str(e)}"
                    metrics_data.append({'instance_id': iid, 'error': error_msg})
                    print(f"ERROR: Failed to collect metrics for {iid}: {e}")

        # Create report with all instances (including stopped ones)
        print("DEBUG: Creating Excel report...")
        excel_file = create_excel_report(metrics_data, running_instances, all_instances, account_number)
        
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        s3_key = f"{S3_KEY_PREFIX}_{timestamp}.xlsx"
        
        print(f"DEBUG: Uploading to S3: {s3_key}")
        s3.put_object(Bucket=S3_BUCKET, Key=s3_key, Body=excel_file)
        
        print("DEBUG: Lambda execution completed successfully")
        return {'statusCode': 200, 'body': f"Report uploaded to s3://{S3_BUCKET}/{s3_key}"}
        
    except Exception as e:
        print(f"ERROR: Lambda execution failed: {str(e)}")
        return {'statusCode': 500, 'body': f"Error: {str(e)}"}